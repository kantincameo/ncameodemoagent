#!/usr/bin/env python3
"""
validate_output.py — Validates a finished 8-object output JSON before
it's shown to the human-in-the-loop.

Usage:
    python3 validate_output.py path/to/output.json [path/to/markers_sheet.xlsx]

Checks:
  - Exactly 8 objects
  - Exactly 2 objects per required group, correct exact group strings
  - Each group has exactly one "highlight" and one "note"
  - Valid UUID v4 for "id", all unique
  - "date" matches YYYY-MM-DDTHH:MM:SSZ
  - No empty "text", "rationale", or "referencesource"
  - Every note's "referencesource" mentions a marker name that actually
    exists in the currently-loaded ground-truth sheet (best-effort check)

Exits non-zero with a printed list of problems if anything fails.
"""
import sys
import os
import json
import re
import uuid

REQUIRED_GROUPS = [
    "Internal Notes",
    "Understanding Your Results",
    "Priority Actions - Next 30 Days",
    "Long-Term Strategy",
]

DATE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2}Z$")

SKILL_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DEFAULT_SHEET = os.path.join(SKILL_ROOT, "references", "BloomCell_Urinalysis.xlsx")


def is_valid_uuid4(s):
    try:
        val = uuid.UUID(s, version=4)
        return str(val) == s.lower()
    except (ValueError, AttributeError):
        return False


def load_marker_names(sheet_path):
    if not os.path.exists(sheet_path):
        return None  # skip that check if sheet unavailable
    try:
        import openpyxl
        wb = openpyxl.load_workbook(sheet_path, data_only=True)
        ws = wb[wb.sheetnames[0]]
        names = []
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            if row[0]:
                names.append(str(row[0]).strip().lower())
        return names
    except Exception:
        return None


def validate(objects, marker_names):
    errors = []

    if len(objects) != 8:
        errors.append(f"Expected exactly 8 objects, got {len(objects)}")

    seen_ids = set()
    group_counts = {g: {"highlight": 0, "note": 0} for g in REQUIRED_GROUPS}

    for i, obj in enumerate(objects):
        prefix = f"Object {i} (id={obj.get('id', '?')}):"

        oid = obj.get("id")
        if not is_valid_uuid4(oid or ""):
            errors.append(f"{prefix} invalid UUID v4: {oid!r}")
        elif oid in seen_ids:
            errors.append(f"{prefix} duplicate id")
        else:
            seen_ids.add(oid)

        date = obj.get("date", "")
        if not DATE_RE.match(date):
            errors.append(f"{prefix} date {date!r} does not match YYYY-MM-DDTHH:MM:SSZ")

        otype = obj.get("type")
        if otype not in ("highlight", "note"):
            errors.append(f"{prefix} invalid type {otype!r}")

        group = obj.get("group")
        if group not in REQUIRED_GROUPS:
            errors.append(f"{prefix} invalid group {group!r}")
        elif otype in ("highlight", "note"):
            group_counts[group][otype] += 1

        for field in ("text", "rationale", "referencesource"):
            if not obj.get(field, "").strip():
                errors.append(f"{prefix} empty field '{field}'")

        if otype == "note" and marker_names is not None:
            ref = (obj.get("referencesource") or "").lower()
            if not any(m in ref for m in marker_names):
                errors.append(
                    f"{prefix} referencesource does not appear to name any marker "
                    f"from the current ground-truth sheet: {obj.get('referencesource')!r}"
                )

    for group, counts in group_counts.items():
        if counts["highlight"] != 1 or counts["note"] != 1:
            errors.append(
                f"Group '{group}' has {counts['highlight']} highlight(s) and "
                f"{counts['note']} note(s), expected exactly 1 of each"
            )

    return errors


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python3 validate_output.py path/to/output.json [path/to/markers_sheet.xlsx]",
              file=sys.stderr)
        sys.exit(1)

    with open(sys.argv[1], "r", encoding="utf-8") as f:
        objects = json.load(f)

    sheet_path = sys.argv[2] if len(sys.argv) > 2 else DEFAULT_SHEET
    marker_names = load_marker_names(sheet_path)

    errors = validate(objects, marker_names)

    if errors:
        print(f"FAILED — {len(errors)} problem(s):")
        for e in errors:
            print(f"  - {e}")
        sys.exit(1)
    else:
        print("PASSED — output is schema-valid and grounded in the current sheet.")
        sys.exit(0)
