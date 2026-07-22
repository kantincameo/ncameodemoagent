#!/usr/bin/env python3
"""
read_markers.py — Loads the BloomCell bloodwork ground-truth sheet.

Usage:
    python3 read_markers.py [path/to/sheet.xlsx]

If no path given, defaults to references/BloomCell_Bloodwork_Markers.xlsx
relative to this skill's root. ALWAYS uses whatever .xlsx is currently in
references/ — this is the "latest file" the human may have updated.

Prints a JSON object: { "<marker name>": {internal_notes, explanation,
plan_30_day, long_term_plan, row_number}, ... }

No internet access, no external lookups — this script only reads the local
bundled file.
"""
import sys
import os
import json

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl --break-system-packages", file=sys.stderr)
    sys.exit(1)

SKILL_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DEFAULT_PATH = os.path.join(SKILL_ROOT, "references", "BloomCell_Bloodwork_Markers.xlsx")

EXPECTED_HEADERS = ["Marker Name", "Internal Notes", "Explanation", "30-Day Plan", "Long-Term Plan"]


def load_markers(path):
    if not os.path.exists(path):
        print(f"ERROR: ground-truth sheet not found at {path}", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(path, data_only=True)
    # Use the first sheet — this skill assumes a single-sheet ground-truth file.
    ws = wb[wb.sheetnames[0]]

    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    if [h.strip() if isinstance(h, str) else h for h in header] != EXPECTED_HEADERS:
        print(f"WARNING: header row is {header!r}, expected {EXPECTED_HEADERS!r}. "
              f"Proceeding by column position (A-E) but verify the sheet hasn't changed shape.",
              file=sys.stderr)

    markers = {}
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), start=2):
        marker_name = row[0]
        if not marker_name or not str(marker_name).strip():
            continue
        markers[str(marker_name).strip()] = {
            "internal_notes": (row[1] or "").strip() if isinstance(row[1], str) else row[1],
            "explanation": (row[2] or "").strip() if isinstance(row[2], str) else row[2],
            "plan_30_day": (row[3] or "").strip() if isinstance(row[3], str) else row[3],
            "long_term_plan": (row[4] or "").strip() if isinstance(row[4], str) else row[4],
            "row_number": row_idx,
        }
    return markers


if __name__ == "__main__":
    path = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_PATH
    markers = load_markers(path)
    print(json.dumps({
        "source_file": os.path.basename(path),
        "marker_count": len(markers),
        "markers": markers,
    }, indent=2, ensure_ascii=False))
